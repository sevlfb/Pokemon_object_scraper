import openpyxl.drawing
import openpyxl.drawing.image
from openpyxl.drawing.picture import PictureFrame
from openpyxl.drawing.xdr import XDRPositiveSize2D
import openpyxl.utils
from openpyxl.styles import Alignment
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.styles import Font, Border, Side, Color, Alignment, PatternFill
from openpyxl.styles import numbers

import urllib3
import io
import math
from libs.utils.bs4_utils import URL_PREFIX
from libs.classes.GameLocations.GameLocationsAbstract import GameLocationsAbstract
from libs.utils.string_utils import get_images_objects_in_string, string_contains_images


"""def get_sheet_name(badge: int, game_locations: GameLocations):
    if int(badge) == 8:
        return "Avant la Ligue"
    elif int(badge) == 9:
        return "Après la Ligue"
    else:
        return f"Avant l'{eval(f"game_enum.ARENE_{int(badge)}")}"
"""    

def get_url_image(http: urllib3.PoolManager, url: str):
    r = http.request('GET', url)
    image_file = io.BytesIO(r.data)
    img = openpyxl.drawing.image.Image(image_file) 
    img.height = 20
    img.width = 20  
    return img


def get_cell_border(iter: int, max_iter: int, col_index:int, max_col: int, ):
    color = Color()
    top = None if iter > 0 else Side(style="thick", color=color)
    bottom = None if iter < max_iter else Side(style="thick", color=color)
    left = Side(style="mediumDashDotDot", color=color) if col_index > 0 else Side(style="thick", color=color)
    right = Side(style="mediumDashDotDot", color=color) if col_index < max_col else Side(style="thick", color=color)
    return Border(top=top, bottom=bottom, left=left, right=right)


def design_place_name(left_cell, right_cell, font_size=14, alignment_ = "left", bgColor="7187F3"):
    bgColor = Color(rgb=bgColor)
    fill = PatternFill("solid", start_color=bgColor)
    font = Font(name="Cascadia Code SemiBold", size=font_size, bold=True)
    alignment = Alignment(horizontal=alignment_)
    left_border = Border(top=Side("thick"), bottom=Side("thick"), left=Side("thick"))
    right_border = Border(top=Side("thick"), bottom=Side("thick"), right=Side("thick"))
    all_border = Border(top=Side("thick"), bottom=Side("thick"), right=Side("thick"), left=Side("thick"))
    if left_cell == right_cell:
        left_cell.border = all_border
    else:
        left_cell.border = left_border
        right_cell.border = right_border
    for cell in [left_cell, right_cell]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
    

def set_or_create_sheet(workbook, badge, game_locations: GameLocationsAbstract):
    if int(badge) == 0:
        sheet = workbook.active
        sheet.title = "Départ"
    else:
        sheet = workbook.create_sheet(game_locations.arena_string(int(badge)))
    return sheet


def set_sheet_cols_dim(sheet, cols):
    IMAGE_SIZE = 3.5
    DEFAULT_FONT_SIZE = 11
    PRETTY_WIDTH = 2
    sheet.column_dimensions[openpyxl.utils.get_column_letter(1)].width = IMAGE_SIZE
    for i, col in enumerate(cols[1:], start=1): 
        excel_col_index = i + 1
        column_cell_sizes = [0 if cell.value is None else len(cell.value) for cell in col]
        # For column B
        if excel_col_index == 2:
            enlarge_factor = lambda cell: len(cell.value)/(cell.value.count("\n")+1) * math.pow(cell.font.__getattr__("size")/(DEFAULT_FONT_SIZE), 2)
            column_cell_sizes += [0 if cell.value is None else enlarge_factor(cell)-IMAGE_SIZE+PRETTY_WIDTH for cell in cols[0]]
        sheet.column_dimensions[openpyxl.utils.get_column_letter(excel_col_index)].width = max(column_cell_sizes)


def set_sheet_rows_dim(sheet, rows, cols=5):
    for i, row in enumerate(rows, start=1): 
        increase_factor = max([cell.count("\n") for cell in row])
        sheet.row_dimensions[openpyxl.utils.get_column_letter(i)].height = increase_factor


def blank_sheet(cols_to_blank):
    for col in cols_to_blank:
        for cell in col:
            if cell.fill.__getattr__("start_color") not in [Color(rgb="7187F3"), Color(rgb="B4E8FA"), Color(rgb="AAB7DA"), Color(rgb="9DBE94")]:
                cell.fill = PatternFill("solid", start_color=Color(rgb="ffffff"))


def chronological_sort(objects_data: dict, game_locations: GameLocationsAbstract):
    ordered_data = []
    for badge in game_locations.locations_order:
        for place in game_locations.locations_order[badge]["places"]:
            if place in objects_data:
                ordered_data.append((place, objects_data[place]))
    return ordered_data


def get_percent_completion(iters: list, lengths: list):
    return int(sum([float(iter)/math.prod(lengths[:i+1]) for i, iter in enumerate(iters)]) * 100)


 
def write_all_data_in_excel(global_object_data, game_locations: GameLocationsAbstract, game:str, object_: str, path="./output"):
    
    """
    data = {
        "badge":
            "Place": {
                "Table Group Title": {
                    "SubPlace": {
                        
                    }
                }
            }
        }
    }
    """
    
    workbook = openpyxl.Workbook()
    http = urllib3.PoolManager()
    EVEN_BACKGROUND_COLOR = Color(rgb="AAB7DA")
    ODD_BACKGROUND_COLOR = Color(rgb="B4E8FA")
    
    for iter_badge, number_badges_obtained in enumerate(global_object_data):
        sheet = set_or_create_sheet(workbook, number_badges_obtained, game_locations)
        row_index = 1
        ordered_places = chronological_sort(global_object_data[number_badges_obtained], game_locations)
        for iter_place, (place_real_name, place_data) in enumerate(ordered_places):
            sheet[f"A{row_index}"].value = place_real_name
            design_place_name(sheet[f"A{row_index}"], sheet[f"B{row_index}"])
            row_index += 1
            for place_sub_group in place_data:
                if place_sub_group != "":
                    sheet[f"A{row_index}"].value = place_sub_group
                    design_place_name(sheet[f"A{row_index}"], sheet[f"A{row_index}"], font_size=11, alignment_ = "center", bgColor="9DBE94")
                    sheet.merge_cells(start_row=row_index, end_row=row_index, start_column=1, end_column=3)
                    row_index += 1
                for sub_place in (sub_group:=place_data[place_sub_group]):
                    if sub_place != "":
                        sheet[f"A{row_index}"].value = sub_place
                        design_place_name(sheet[f"A{row_index}"], sheet[f"A{row_index}"], font_size=11, alignment_ = "center")
                        sheet.merge_cells(start_row=row_index, end_row=row_index, start_column=1, end_column=3)
                        row_index += 1
                    for object_iter, object_data_row in enumerate(temp_ := sub_group[sub_place]):
                        #print(number_badges_obtained, place_real_name, place_sub_group, sub_place, object_data_row, sep = " -> ")
                        print(f"""{get_percent_completion([iter_badge, iter_place],
                            [len(global_object_data), len(global_object_data[number_badges_obtained])])}%""", end="\r")
                        for field_index, object in enumerate(object_data_row):
                            cell = sheet.cell(row_index, field_index+1)
                            border = get_cell_border(object_iter, len(temp_)-1, field_index, len(object_data_row)-1)
                            cell.border = border
                            bgColor =  ODD_BACKGROUND_COLOR if object_iter%2==0 else EVEN_BACKGROUND_COLOR
                            cell.fill = PatternFill("solid", start_color=bgColor)
                            IMAGE_INDEX = 0
                            cell = sheet.cell(row_index, field_index+1)
                            cell.alignment = Alignment(horizontal='left',
                                                       vertical='top',
                                                       wrap_text=True)
                            try:
                                cell.value = "  "+str(object_data_row[field_index])
                            except:
                                # take above value
                                cell.value = sheet.cell(row_index-1, field_index).value
                            cell.number_format = numbers.FORMAT_TEXT
                            if "%" in cell.value and "\n" not in cell.value:
                                cell.value = cell.value.strip()
                                cell.number_format = numbers.FORMAT_PERCENTAGE
                            if "\n" in cell.value:
                                cell.font = Font(size=int(11/(cell.value.count("\n")+1)*1.5))
                            if string_contains_images(cell.value):
                                #print(f"{URL_PREFIX}{object_data_row[field_index]}")
                                images = get_images_objects_in_string(str(object_data_row[field_index]))
                                for image_iter, image in enumerate(images):
                                    EMU_PER_PIXEL = 9525
                                    cell.value = cell.value.replace(image, "")
                                    cell.value += " "*4
                                    col_index = field_index+1 if field_index == 1 else field_index 
                                    colOffset = EMU_PER_PIXEL*25 if field_index == 1 else 0
                                    marker = AnchorMarker(col=col_index, row=row_index-1, 
                                                          colOff=(-EMU_PER_PIXEL*20*image_iter-colOffset), rowOff=0)
                                    img = get_url_image(http, f"{URL_PREFIX}{image}")
                                    ext_width = img.width * EMU_PER_PIXEL
                                    ext_height = img.height * EMU_PER_PIXEL
                                    ext = XDRPositiveSize2D(ext_width, ext_height)
                                    img.anchor = OneCellAnchor(_from=marker, ext=ext)
                                    sheet.add_image(img, f'{cell.column_letter}{cell.row}')
                                    img.anchor = OneCellAnchor(_from=marker, ext=ext)

                                #cell.value = f"=_xlfn.IMAGE(LIEN_HYPERTEXTE(\"{URL_PREFIX}{object[i]}\"))"
                        row_index += 1
                row_index += 1
                
        row_index += 2
        
        cols = list(sheet.iter_cols(1,5))
        set_sheet_cols_dim(sheet, cols)
        
        cols_to_blank = sheet.iter_cols(1, 26, 1, sheet.max_row + 15)
        blank_sheet(cols_to_blank)

    if game is None or object_ is None:
        workbook.save(f"{path}/Scrapped_objects.xlsx")
    else:
        workbook.save(f"{path}/{game}_{object_}.xlsx")